import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ingredientes"

ws.append(["nombre", "descripcion", "es_alergeno"])
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 15

comunes = [
    ("Sal", "Cloruro de sodio, condimento básico", "FALSE"),
    ("Pimienta negra", "Especia molida de uso universal", "FALSE"),
    ("Ajo", "Bulbo aromático fresco o deshidratado", "FALSE"),
    ("Cebolla", "Hortaliza de uso básico en cocina", "FALSE"),
    ("Tomate", "Fruto rojo maduro, base de salsas", "FALSE"),
    ("Aceite de oliva", "Aceite vegetal de primera presión en frío", "FALSE"),
    ("Aceite de girasol", "Aceite vegetal refinado para frituras", "FALSE"),
    ("Azúcar", "Sacarosa refinada blanca", "FALSE"),
    ("Azúcar morena", "Azúcar de caña sin refinar", "FALSE"),
    ("Sal marina", "Sal obtenida por evaporación de agua de mar", "FALSE"),
    ("Vinagre de manzana", "Vinagre obtenido de la fermentación de manzana", "FALSE"),
    ("Vinagre blanco", "Vinagre de alcohol destilado", "FALSE"),
    ("Pimentón dulce", "Pimiento rojo seco y molido", "FALSE"),
    ("Pimentón picante", "Pimiento seco picante molido", "FALSE"),
    ("Orégano", "Hierba aromática seca", "FALSE"),
    ("Albahaca", "Hierba fresca o seca de aroma intenso", "FALSE"),
    ("Perejil", "Hierba fresca de uso decorativo y culinario", "FALSE"),
    ("Romero", "Hierba aromática para carnes y asados", "FALSE"),
    ("Tomillo", "Hierba aromática de aroma herbal", "FALSE"),
    ("Laurel", "Hoja aromática para guisos y caldos", "FALSE"),
    ("Cilantro", "Hierba fresca de sabor intenso", "FALSE"),
    ("Cúrcuma", "Especia de color amarillo intenso", "FALSE"),
    ("Comino", "Semilla aromática de uso en especias", "FALSE"),
    ("Jengibre", "Raíz fresca o en polvo de sabor picante", "FALSE"),
    ("Canela", "Corteza aromática dulce-especiada", "FALSE"),
    ("Vainilla", "Extracto o esencia de vaina de vainilla", "FALSE"),
    ("Cacao en polvo", "Polvo obtenido del cacao sin azúcar", "FALSE"),
    ("Levadura seca", "Levadura de panadería deshidratada", "FALSE"),
    ("Bicarbonato de sodio", "Agente leudante para repostería", "FALSE"),
    ("Polvo de hornear", "Mezcla leudante para masas", "FALSE"),
    ("Maicena", "Almidón de maíz espesante", "FALSE"),
    ("Arroz blanco", "Cereal base hervido o al vapor", "FALSE"),
    ("Arroz integral", "Cereal de grano entero rico en fibra", "FALSE"),
    ("Quinoa", "Pseudocereal andino rico en proteínas", "FALSE"),
    ("Lentejas", "Legumbre seca de cocción rápida", "FALSE"),
    ("Garbanzos", "Legumbre seca versátil", "FALSE"),
    ("Porotos negros", "Legumbre de color oscuro y sabor terroso", "FALSE"),
    ("Porotos blancos", "Legumbre suave ideal para guisos", "FALSE"),
    ("Arvejas", "Legumbre verde fresca o congelada", "FALSE"),
    ("Choclo", "Maíz tierno en grano", "FALSE"),
    ("Zanahoria", "Raíz vegetal naranja rica en betacaroteno", "FALSE"),
    ("Apio", "Tallo vegetal crujiente y aromático", "FALSE"),
    ("Puerro", "Vegetal de la familia de la cebolla", "FALSE"),
    ("Calabaza", "Vegetal de pulpa naranja y dulce", "FALSE"),
    ("Zapallo", "Hortaliza de pulpa amarilla o naranja", "FALSE"),
    ("Berenjena", "Vegetal morado de sabor suave", "FALSE"),
    ("Zucchini", "Calabacín verde de textura suave", "FALSE"),
    ("Pimiento rojo", "Fruto dulce de color rojo", "FALSE"),
    ("Pimiento verde", "Fruto de sabor más amargo que el rojo", "FALSE"),
    ("Pimiento amarillo", "Fruto dulce de color amarillo", "FALSE"),
    ("Lechuga", "Hoja verde crujiente para ensaladas", "FALSE"),
    ("Rúcula", "Hoja verde amarga para ensaladas", "FALSE"),
    ("Espinaca", "Hoja verde rica en hierro", "FALSE"),
    ("Acelga", "Vegetal de hojas verdes y tallo blanco", "FALSE"),
    ("Brócoli", "Vegetal crucífero verde rico en vitaminas", "FALSE"),
    ("Coliflor", "Vegetal crucífero blanco", "FALSE"),
    ("Repollo", "Vegetal de hojas apretadas", "FALSE"),
    ("Hongos portobello", "Hongo grande de sabor intenso", "FALSE"),
    ("Champiñones", "Hongo blanco de uso general", "FALSE"),
    ("Cebolla de verdeo", "Cebolla tierna con tallo verde", "FALSE"),
    ("Ají molido", "Condimento picante en polvo", "FALSE"),
    ("Mostaza en pasta", "Condimento cremoso de semilla de mostaza", "FALSE"),
    ("Ketchup", "Salsa de tomate dulce y especiada", "FALSE"),
    ("Salsa worcestershire", "Salsa inglesa umami de sabor complejo", "FALSE"),
    ("Salsa de soja", "Condimento salado fermentado", "FALSE"),
    ("Aceite de coco", "Grasa vegetal extraída del coco", "FALSE"),
    ("Limón", "Cítrico ácido para condimentar", "FALSE"),
    ("Lima", "Cítrico pequeño de sabor intenso", "FALSE"),
    ("Naranja", "Cítrico dulce para jugos y salsas", "FALSE"),
    ("Miel", "Endulzante natural producido por abejas", "FALSE"),
    ("Stevia", "Endulzante natural sin calorías", "FALSE"),
    ("Caldo de verduras", "Caldo base vegetal para sopas", "FALSE"),
    ("Caldo de pollo", "Caldo base de pollo para guisos", "FALSE"),
    ("Tomate triturado", "Pulpa de tomate procesada en lata", "FALSE"),
    ("Puré de tomate", "Concentrado de tomate sin semillas", "FALSE"),
    ("Aceitunas negras", "Aceitunas curadas en salmuera", "FALSE"),
    ("Aceitunas verdes", "Aceitunas verdes con o sin hueso", "FALSE"),
    ("Alcaparras", "Brotes en vinagre de uso culinario", "FALSE"),
    ("Pepinillos en vinagre", "Pepinos pequeños encurtidos", "FALSE"),
    ("Ananá", "Fruta tropical dulce y ácida", "FALSE"),
    ("Mango", "Fruta tropical de pulpa amarilla", "FALSE"),
    ("Banana", "Fruta tropical rica en potasio", "FALSE"),
    ("Frutilla", "Fruto rojo dulce y ácido", "FALSE"),
    ("Arándanos", "Frutos pequeños azules antioxidantes", "FALSE"),
    ("Frambuesas", "Frutos rojos pequeños y aromáticos", "FALSE"),
    ("Durazno", "Fruta de carozo de pulpa suave", "FALSE"),
    ("Pera", "Fruta de sabor dulce y textura suave", "FALSE"),
    ("Manzana", "Fruta crujiente de múltiples variedades", "FALSE"),
    ("Uvas", "Frutos en racimo de sabor dulce", "FALSE"),
    ("Coco rallado", "Pulpa de coco deshidratada y rallada", "FALSE"),
    ("Extracto de malta", "Endulzante espeso derivado de la cebada", "FALSE"),
    ("Agua mineral", "Agua con minerales naturales", "FALSE"),
    ("Jarabe de arce", "Endulzante natural de savia de arce", "FALSE"),
    ("Agar agar", "Gelificante vegetal derivado de algas", "FALSE"),
    ("Gelatina sin sabor", "Proteína colagena para gelificación", "FALSE"),
    ("Cebolla en polvo", "Cebolla deshidratada y molida", "FALSE"),
    ("Ajo en polvo", "Ajo deshidratado y molido", "FALSE"),
    ("Hierbas provenzales", "Mezcla de hierbas secas mediterráneas", "FALSE"),
    ("Curry en polvo", "Mezcla de especias de origen indio", "FALSE"),
    ("Paprika ahumada", "Pimentón ahumado de sabor intenso", "FALSE"),
    ("Pimienta blanca", "Especia de sabor suave y aroma delicado", "FALSE"),
    ("Cardamomo", "Especia aromática de vainas verdes", "FALSE"),
    ("Clavo de olor", "Especia de aroma intenso y dulce", "FALSE"),
    ("Nuez moscada", "Semilla aromática para repostería y salsas", "FALSE"),
    ("Anís estrellado", "Especia de sabor anisado en forma de estrella", "FALSE"),
    ("Chipotle", "Jalapeño ahumado de sabor picante y ahumado", "FALSE"),
]

alergenos = [
    ("Leche", "Leche de vaca entera o semidescremada", "TRUE"),
    ("Huevo", "Huevo de gallina fresco", "TRUE"),
    ("Trigo", "Cereal con gluten base del pan y pasta", "TRUE"),
    ("Gluten", "Proteína presente en trigo, cebada y centeno", "TRUE"),
    ("Maní", "Legumbre de alto potencial alergénico", "TRUE"),
    ("Almendras", "Fruto seco de la familia de las rosáceas", "TRUE"),
    ("Nueces", "Fruto seco de árbol con cáscara dura", "TRUE"),
    ("Avellanas", "Fruto seco pequeño de sabor suave", "TRUE"),
    ("Anacardos", "Fruto seco en forma de riñón", "TRUE"),
    ("Pistachos", "Fruto seco verde de cáscara dura", "TRUE"),
    ("Nueces de Brasil", "Fruto seco grande de América del Sur", "TRUE"),
    ("Nueces de macadamia", "Fruto seco cremoso de origen hawaiano", "TRUE"),
    ("Pecanas", "Fruto seco de la familia del nogal", "TRUE"),
    ("Soja", "Legumbre de alta proteína y usos múltiples", "TRUE"),
    ("Sésamo", "Semilla oleaginosa de alto potencial alergénico", "TRUE"),
    ("Mostaza (semilla)", "Semilla de mostaza con proteínas alergénicas", "TRUE"),
    ("Apio (semilla)", "Semilla de apio con componentes alergénicos", "TRUE"),
    ("Sulfitos", "Conservantes usados en vinos y frutas secas", "TRUE"),
    ("Moluscos", "Mariscos como mejillones, calamares y pulpo", "TRUE"),
    ("Crustáceos", "Langostinos, cangrejos y camarones", "TRUE"),
    ("Pescado", "Proteínas de pescado como salmón, atún o merluza", "TRUE"),
    ("Lupino", "Legumbre usada como sustituto de soja", "TRUE"),
    ("Cebada", "Cereal con gluten usado en maltas y cervezas", "TRUE"),
    ("Centeno", "Cereal con gluten de sabor intenso", "TRUE"),
    ("Avena", "Cereal con trazas de gluten", "TRUE"),
    ("Espelta", "Variedad ancestral de trigo con gluten", "TRUE"),
    ("Kamut", "Variedad de trigo duro ancestral con gluten", "TRUE"),
    ("Polen de abeja", "Sustancia recolectada por abejas potencialmente alergénica", "TRUE"),
    ("Látex (proteínas)", "Proteínas de látex con reactividad cruzada", "TRUE"),
    ("Kiwi", "Fruta de alto potencial alergénico en algunos individuos", "TRUE"),
]

for row in comunes + alergenos:
    ws.append(row)

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    if i % 2 == 0:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="EBF3FB")

out = r"C:\Users\Acel\Desktop\ingredientes_foodstore.xlsx"
wb.save(out)
total = ws.max_row - 1
alerg_count = sum(1 for r in comunes + alergenos if r[2] == "TRUE")
print(f"Excel creado: {out}")
print(f"Total ingredientes: {total} ({alerg_count} alérgenos)")
